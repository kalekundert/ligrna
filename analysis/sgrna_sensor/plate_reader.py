#!/usr/bin/env python3

import re, nonstdlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

class BiotekExperiment:
    # This file format sucks.

    class Parser:

        def __init__(self, expt):
            self.expt = expt

        def parse_row(self, row):
            raise NotImplementedError

        def finish(self):
            pass

    class NullParser(Parser):

        def parse_row(self, row):
            pass

    class HeaderParser(Parser):

        header_keys = {
                'Software Version',
                'Experiment File Path:',
                'Protocol File Path:',
                'Plate Number',
                'Date',
                'Time',
                'Reader Type:',
                'Reader Serial Number:',
                'Reading Type',
                'Procedure Details',
                'Plate Type',
        }

        def parse_row(self, row):
            key = row[0].value
            value = row[1].value

            if key in self.header_keys:
                attr = nonstdlib.slugify(key)
                setattr(self.expt, attr, value)
            
    class ReadParser(Parser):

        def parse_row(self, row):
            indent = row[0].value
            key = row[1].value
            read_pattern = re.compile('(?:(?:Blank )?Read \d:)?(\d+)')

            if indent is not None:
                return
            if key is None:
                return
            if key == 'Well':
                self.wells = [x.value for x in row[2:] if x.value]
                return
            
            read_match = read_pattern.match(str(key))
            if read_match:
                wavelength = int(read_match.group(1))
                reads = {
                        k: x.value
                        for k,x in zip(self.wells, row[2:])
                        if x.value and x.value != '?????'
                }
                self.expt.reads[wavelength] = reads

    class KineticParser(Parser):

        def parse_row(self, row):
            title = row[0].value
            key = row[1].value

            if title is not None:
                self.wavelength = int(title.split(':')[-1])
                self.kinetic = []
                self.prev_minutes = 0

            if key == 'Time':
                self.header = ['minutes'] + [x.value for x in row[2:]]

            elif key is not None:
                for i, cell in enumerate(row[3:], 2):
                    # Update the minutes accounting for the fact that openpyxl 
                    # wraps hours after 1 day.  Fucking stupid library.
                    minutes = 60 * key.hour + key.minute
                    while self.prev_minutes > minutes:
                        minutes += 60 * 24
                    self.prev_minutes = minutes

                    entry = {
                            'well': self.header[i],
                            'temperature': row[2].value,
                            'minutes': minutes,
                            'wavelength': self.wavelength,
                            'read': cell.value,
                    }
                    self.kinetic.append(entry)

        def finish(self):
            df = pd.DataFrame(self.kinetic)
            df[df == '?????'] = np.nan
            #df = df.dropna(axis='columns', how='all')
            self.expt.kinetic[self.wavelength] = df

    def __init__(self, path):
        self.reads = {}
        self.kinetic = {}

        wb = load_workbook(path)
        ws = wb.active
        parser = self.HeaderParser(self)
        kinetic_pattern = re.compile('^.*:\d+')

        for row in ws.rows:
            key = row[0].value
            subkey = row[1].value

            if key == 'Results':
                parser.finish()
                parser = self.ReadParser(self)

            if isinstance(key, str) and kinetic_pattern.match(key):
                parser.finish()
                parser = self.KineticParser(self)

            parser.parse_row(row)

        parser.finish()

    def __str__(self):
        from pprint import pformat
        lines = ['BetaGalExperiment']
        for attr, value in self.__dict__.items():
            if attr.startswith('_'):
                continue
            if (attr == 'kinetic' and self.kinetic) or \
               (attr == 'reads' and self.reads):
                lines += [f'  {attr}:', f'{pformat(value)}']
            else:
                lines += [f'  {attr + ":":25s} {value}']
        return '\n'.join(lines)

class PerkinElmerExperiment:

    def __init__(self, path):

        def find_list_sheet(wb):
            for name in wb.sheetnames:
                if name.startswith('List'):
                    return wb[name]

        def min_from_days(days):
            return days * 24 * 60

        wb = load_workbook(path)
        ws = find_list_sheet(wb)

        # The time column is in units of days (?).
        col_titles = []
        time_cols = []
        data = {}

        for i, row in enumerate(ws.rows):
            cells = [x.value for x in row if x.value]

            if i == 0:
                col_titles = cells
                plate_col = cells.index('Plate')
                well_col = cells.index('Well')
                repeat_col = cells.index('Repeat')
                type_col = cells.index('Type')
                time_cols = [i for i,x in enumerate(cells) if x == 'Time']

            else:
                for time_col in time_cols:
                    datum_col = time_col + 1
                    datum_name = nonstdlib.slugify(col_titles[datum_col])
                    datum = { #
                            'plate':    cells[plate_col],
                            'well':     cells[well_col],
                            'repeat':   cells[repeat_col],
                            'type':     cells[type_col],
                            'minutes':  min_from_days(cells[time_col]),
                            datum_name: cells[datum_col],
                    }
                    data.setdefault(datum_name, []).append(datum)

        self.reads = {k: pd.DataFrame(v) for k,v in data.items()}

