#!/usr/bin/env python3

import re
import shutil
import functools
import contextlib
import numpy as np
import pandas as pd
from pathlib import Path
from pprint import pprint

DISPLAY = ['spacer', 'design', 'ligand', 'band', 'pixels']
INHERITABLE_COLS = 'spacer', 'design', 'ligand'

pd.options.display.width = shutil.get_terminal_size().columns

def standardize_df(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        df = f(*args, **kwargs)
        df = natsort_spacers(df.reset_index())
        if 'index' in df: del df['index']
        return df
    return wrapper


def load_cleavage_data_from_xlsx_dir(dir, **kw):
    dir = Path(dir)
    dfs = [load_cleavage_data_from_xlsx(p, **kw)
            for p in sorted(dir.glob('*.xlsx'))]
    return pd.concat(dfs, ignore_index=True)

def load_cleavage_data_from_xlsx(path, drop_rejects=True, inheritable_cols=INHERITABLE_COLS):
    try:
        from openpyxl import load_workbook
        book = load_workbook(path)
        sheet = book['data']

        # Start by parsing the column headers.  Not all the spreadsheets will 
        # have the same data in the same columns, so we need to figure out 
        # which columns are present before parsing the rows.

        headers = []
        cols = []
        is_col_empty = lambda j: sheet.cell(row=1, column=j).value is None
        j = 1  # indexing is from 1.

        while not is_col_empty(j):
            title = sheet.cell(row=1, column=j).value

            # Attempt to remove weird characters from the column titles, 
            # including parenthesized unit labels and punctuation.  This allows 
            # downstream code to use pull data out of the data frame using the 
            # nicer attribute syntax.

            slug = title.lower()
            slug = re.sub('\(.*\)', '', slug)
            slug = re.sub('\W', '', slug)

            ignore = 'cleaved', 'change', 'notes'
            if slug not in ignore:
                headers.append(slug)
                cols.append(j)

            j += 1

        # Parse the data from each row.  For "inheritable" columns ('expt', 
        # 'spacer', 'design', and 'ligand' by default), fill in missing values 
        # from previous rows.
        
        rows = []
        previous_row = []
        is_row_empty = lambda i: not any([
            sheet.cell(row=i, column=j).value for j in cols])
        i = 2  # Skip the header (indexing is from 1).

        def parse_cell(i, j, j0): #
            x = sheet.cell(row=i, column=j).value
            if x is None and headers[j0] in INHERITABLE_COLS:
                x = previous_row[j0]
            return x

        while not is_row_empty(i):
            row = [parse_cell(i, j, j0) for j0, j in enumerate(cols)]
            rows.append(row)
            previous_row = row
            i += 1

        # Convert the parsed data to a pandas data frame, then clean it up a 
        # little bit and check for data entry errors.

        df = pd.DataFrame(rows, columns=headers)

        df['expt'] = str(path)
        df['date'] = pd.Timestamp(re.search('[0-9]{8}', str(path)).group(0))
        df['order'] = range(len(df))

        df = drop_rejected_data(df)
        df = sanitize_data(df)

    except:
        msg = f"Unexpected error while parsing '{path}'"
        raise ValueError(msg)

    check_for_errors(df, path)
    return df
    

def natsort_spacers(df):
    from natsort import order_by_index, index_natsorted
    return df.reindex(
            index=order_by_index(df.index, index_natsorted(df.spacer)),
    )

def sort_spacers_by_activity(df, designs):
    activities = np.vstack([
        df[df.design == x].mean_change.values
        for x in designs
    ])
    ranks = np.argsort(np.argsort(abs(activities)))
    summed_ranks = np.sum(ranks, axis=0)
    most_to_least_activity = np.argsort(summed_ranks)[::-1]

    spacers = df['spacer'].unique()
    return spacers[most_to_least_activity]

def sort_by_activity(df, designs):
    order = sort_spacers_by_activity(df, designs)
    df['spacer'] = pd.Categorical(df['spacer'], order)
    return df.sort_values('spacer')

def drop_rejected_data(df):
    if 'reject' not in df:
        return df

    cols = {'expt', 'spacer', 'design'} & set(df.columns)
    df = df.groupby(list(cols)).\
            filter(lambda x: x.reject.isnull().all())
    del df['reject']
    return df

def sanitize_data(df):
    """
    Iron out some idiosyncrasies in the data input.  For example, initially I 
    used '+' and '−' to indicate apo and holo reactions, but later I decided 
    that concentrations (10,000 μM) were more informative.
    """
    if 'spacer' not in df:
        df['spacer'] = 'aavs'

    df.loc[df.ligand == '−', 'ligand'] = 0      # df.loc[<row>, <col>]
    df.loc[df.ligand == '+', 'ligand'] = 10000
    df.loc[df.design == 'rxb 11', 'design'] = 'rxb 11,1'

    return df

def check_for_errors(df, path):
    # Make sure I didn't leave in any pixels I meant to set to 0.
    if (df.pixels < 0).any():
        msg = f"Found pixels < 0 in '{path}', that shouldn't happen..."
        raise ValueError(msg)
    if (df.pixels > 50_000).any():
        msg = f"Found pixels > 50,000 in '{path}', did you forget to zero them?"
        raise ValueError(msg)

    # Make sure all the data have spacers
    if df.spacer.apply(lambda x: x.strip()).isnull().any():
        msg = f"Found data without a spacer in '{path}'"
        raise ValueError(msg)

    # Make sure none of the designs are misspelled.
    from .usage import from_name
    unexpected_names = []
    for name in set(df.design):
        # I don't give the full name for zipper designs, so ignore those.
        if name[0] == 'z' or name[:2] == 'id':
            continue
        try: from_name(name)
        except (ValueError, IndexError): unexpected_names.append(name)
    if unexpected_names:
        msg = f"Found the following unexpected designs in '{path}': {', '.join(str(x) for x in sorted(unexpected_names))}"
        raise ValueError(msg)

    # Make sure none of the bands are misspelled
    expected_bands = {4000, 2000, 500, 350}
    unexpected_bands = set(df.band) - expected_bands
    if unexpected_bands:
        msg = f"Found the following unexpected bands in '{path}': {', '.join(str(x) for x in sorted(unexpected_bands))}"
        raise ValueError(msg)


def calc_percent_cut(df):
    if 'percent_cut' in df: return df
    return df.\
            groupby(['spacer', 'design', 'expt', 'ligand']).\
            apply(percent_cut)

def percent_cut(group):
    if set(group.band) == {500, 350}:
        uncut_px = 500 * group[group.band == 500].pixels.iat[0]
        cut_px = 350 * group[group.band == 350].pixels.iat[0]
    elif set(group.band) == {4000, 2000}:
        uncut_px = group[group.band == 4000].pixels.iat[0]
        cut_px = group[group.band == 2000].pixels.iat[0]
    else:
        raise ValueError(f'Unexpected bands:\n{group}')

    return pd.Series({
        'percent_cut': cut_px / (uncut_px + cut_px),
        'date': min(group.date),
        'order': min(group.order),
    })

def calc_percent_change(df):
    if 'percent_change' in df: return df
    return calc_percent_cut(df).\
            groupby(['spacer', 'design', 'expt']).\
            apply(percent_change)

def percent_change(group):
    group = group.reset_index()  # MultiIndex confuses me...so just nuke it...
    apo_percent = group[group.ligand == 0].percent_cut.iat[0]
    holo_percent = group[group.ligand == 10000].percent_cut.iat[0]
    return pd.Series({
        'apo_percent': apo_percent,
        'holo_percent': holo_percent,
        'percent_change': holo_percent - apo_percent,
        'date': min(group.date),
        'order': min(group.order),
    })

def calc_mean_change(df):
    if 'mean_change' in df: return df
    return calc_percent_change(df).\
            groupby(['spacer', 'design']).\
            apply(mean_change).reset_index()

def mean_change(group):
    earliest = group.date.idxmin()

    def most_extreme(x): #
        a, b = min(x), max(x)
        return a if abs(a) > abs(b) else b

    def least_extreme(x): #
        a, b = min(x), max(x)
        return a if abs(a) < abs(b) else b

    return pd.Series({
        'earliest_date': group.date.loc[earliest],
        'earliest_order': group.order.loc[earliest],

        'mean_change': group.percent_change.mean(),
        'min_change': least_extreme(group.percent_change),
        'max_change': most_extreme(group.percent_change),
        'std_change': group.percent_change.std(),

        'mean_apo': group.apo_percent.mean(),
        'min_apo': least_extreme(group.apo_percent),
        'max_apo': most_extreme(group.apo_percent),
        'std_apo': group.apo_percent.std(),

        'mean_holo': group.holo_percent.mean(),
        'min_holo': least_extreme(group.holo_percent),
        'max_holo': most_extreme(group.holo_percent),
        'std_holo': group.holo_percent.std(),
        
        'num_replicates': len(group),
    })


@contextlib.contextmanager
def plot_or_savefig(output_path=None, substitution_path=None, inkscape_svg=False):
    """
    Either open the plot in the default matplotlib GUI or export the plot to a 
    file, depending on whether or not an output path is given.  If an output 
    path is given and it contains dollar signs ('$'), they will be replaced 
    with the given substitution path.
    """
    import os, sys, subprocess, matplotlib.pyplot as plt
    from pathlib import Path
    from tempfile import NamedTemporaryFile

    # Decide how we are going to display the plot.  We have to do this before 
    # anything is plotted, because if we want to 

    if output_path is None:
        fate = 'gui'
    elif output_path is False:
        fate = 'none'
    elif output_path == 'lpr':
        fate = 'print'
    else:
        fate = 'save'

    # We have to decide whether or not to fork before plotting anything, 
    # otherwise X11 will complain, and we only want to fork if we'll end up 
    # showing the GUI.  So first we calculate the output path, then we either 
    # fork or don't, then we yield to let the caller plot everything, then we 
    # either display the GUI or save the figure to a file.

    if fate == 'gui' and os.fork():
        sys.exit()

    yield

    if fate == 'print':
        temp_file = NamedTemporaryFile(prefix='sgrna_sensor_', suffix='.ps')
        plt.savefig(temp_file.name, dpi=300)
        # Add ['-o', 'number-up=4'] to get plots that will fit in a paper lab 
        # notebook.
        subprocess.call(['lpr', temp_file.name])

    if fate == 'save':
        if substitution_path:
            output_path = output_path.replace('$', Path(substitution_path).stem)

        if not inkscape_svg:
            plt.savefig(output_path, dpi=300)

        else:
            from tempfile import NamedTemporaryFile
            prefix = Path(output_path).stem + '_'
            with NamedTemporaryFile(prefix=prefix, suffix='.pdf') as pdf:
                plt.savefig(pdf.name, dpi=300)
                subprocess.call([
                        'inkscape', 
                        '--without-gui',
                        '--file', pdf.name,
                        '--export-plain-svg', output_path,
                ])

    if fate == 'gui':
        plt.gcf().canvas.set_window_title(' '.join(sys.argv))
        plt.show()


